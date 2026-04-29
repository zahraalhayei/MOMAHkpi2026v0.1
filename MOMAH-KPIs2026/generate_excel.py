import os
import re
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── colours ──────────────────────────────────────────────────────────────────
DARK_GREEN   = "00563F"
LIGHT_GREEN  = "E8F5E9"
WHITE        = "FFFFFF"

HDR_FILL  = PatternFill("solid", start_color=DARK_GREEN, end_color=DARK_GREEN)
ALT_FILL  = PatternFill("solid", start_color=LIGHT_GREEN, end_color=LIGHT_GREEN)
WHITE_FILL = PatternFill("solid", start_color=WHITE,      end_color=WHITE)

HDR_FONT   = Font(name="Arial", bold=True, color=WHITE,      size=11)
TITLE_FONT = Font(name="Arial", bold=True, color=DARK_GREEN, size=14)
DATA_FONT  = Font(name="Arial", size=10)
SUMM_HDR_FONT = Font(name="Arial", bold=True, color=WHITE, size=11)

RTL_CENTER = Alignment(horizontal="center", vertical="center",
                        wrap_text=True,  readingOrder=2)
RTL_RIGHT  = Alignment(horizontal="right",  vertical="center",
                        wrap_text=True,  readingOrder=2)
RTL_LEFT   = Alignment(horizontal="left",   vertical="center",
                        wrap_text=True,  readingOrder=2)

THIN = Side(style="thin", color="CCCCCC")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ── parse JS file ─────────────────────────────────────────────────────────────
def extract_js_objects(js_text):
    """Strip JS syntax so we can hand individual array contents to json.loads."""

    def find_matching_brace(s, start, open_ch, close_ch):
        depth, i = 0, start
        while i < len(s):
            if s[i] == open_ch:
                depth += 1
            elif s[i] == close_ch:
                depth -= 1
                if depth == 0:
                    return i
            i += 1
        return -1

    def js_to_json(raw):
        # remove single-line comments
        raw = re.sub(r'//[^\n]*', '', raw)
        # remove trailing commas before } or ]
        raw = re.sub(r',\s*([}\]])', r'\1', raw)
        # convert single-quoted strings to double-quoted
        # handle escaped single quotes inside
        def replace_sq(m):
            inner = m.group(1).replace('\\"', '__DQUOTE__')
            inner = inner.replace('"', '\\"')
            inner = inner.replace("\'", "'")
            inner = inner.replace('__DQUOTE__', '\\"')
            return '"' + inner + '"'
        raw = re.sub(r"'((?:[^'\\]|\\.)*)'", replace_sq, raw)
        # quote unquoted keys (identifier chars before colon)
        raw = re.sub(r'(?<=[{,])\s*([A-Za-z_$؀-ۿݐ-ݿﭐ-﷿ﹰ-﻿][A-Za-z0-9_$؀-ۿݐ-ݿﭐ-﷿ﹰ-﻿]*)\s*:', r' "\1":', raw)
        # JS booleans / null are already valid JSON
        return raw

    def get_array(text, key):
        pattern = re.compile(r'["\']?' + re.escape(key) + r'["\']?\s*:\s*\[', re.DOTALL)
        m = pattern.search(text)
        if not m:
            return []
        start = m.end() - 1           # position of '['
        end   = find_matching_brace(text, start, '[', ']')
        raw   = text[start:end+1]
        raw   = js_to_json(raw)
        try:
            return json.loads(raw)
        except json.JSONDecodeError as e:
            print(f"JSON parse error for key '{key}': {e}")
            print("Raw snippet:", raw[:300])
            return []

    mock_m = re.search(r'const\s+MOCK_DATA\s*=\s*\{', js_text)
    extra_m = re.search(r'const\s+EXTRA_DATA\s*=\s*\{', js_text)

    mock_start  = mock_m.start()  if mock_m  else 0
    extra_start = extra_m.start() if extra_m else len(js_text)

    mock_text  = js_text[mock_start:extra_start]
    extra_text = js_text[extra_start:]

    mock = {
        "bi":       get_array(mock_text,  "bi"),
        "ops":      get_array(mock_text,  "ops"),
        "strategy": get_array(mock_text,  "strategy"),
    }
    extra = {
        "performance":   get_array(extra_text, "performance"),
        "workplan":      get_array(extra_text, "workplan"),
        "interventions": get_array(extra_text, "interventions"),
        "reports":       get_array(extra_text, "reports"),
    }
    return mock, extra


# ── helpers ───────────────────────────────────────────────────────────────────
def write_title(ws, title_text, ncols):
    ws.row_dimensions[1].height = 32
    cell = ws.cell(row=1, column=1, value=title_text)
    cell.font      = TITLE_FONT
    cell.alignment = RTL_CENTER
    cell.fill      = WHITE_FILL
    if ncols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)


def write_header(ws, row, headers):
    ws.row_dimensions[row].height = 24
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font      = HDR_FONT
        c.fill      = HDR_FILL
        c.alignment = RTL_CENTER
        c.border    = CELL_BORDER


def write_data_row(ws, row, values, alt):
    fill = ALT_FILL if alt else WHITE_FILL
    ws.row_dimensions[row].height = 18
    for col, v in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.font      = DATA_FONT
        c.fill      = fill
        c.alignment = RTL_RIGHT
        c.border    = CELL_BORDER


def autofit(ws, min_w=15, max_w=40):
    for col_cells in ws.columns:
        best = min_w
        for cell in col_cells:
            if cell.value:
                # rough char-width estimate (Arabic chars wider)
                length = len(str(cell.value))
                best = max(best, min(length * 1.3, max_w))
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = best


def set_sheet_style(ws, tab_color=DARK_GREEN):
    ws.sheet_properties.tabColor = tab_color
    ws.sheet_view.rightToLeft    = True


def add_data_sheet(wb, sheet_name, rows, title_suffix="وزارة البلديات والإسكان | 2026"):
    ws = wb.create_sheet(title=sheet_name)
    set_sheet_style(ws)

    if not rows:
        ws.cell(row=1, column=1, value="لا توجد بيانات")
        return ws

    headers = list(rows[0].keys())
    ncols   = len(headers)

    write_title(ws, f"{sheet_name} - {title_suffix}", ncols)
    write_header(ws, 2, headers)

    for i, row in enumerate(rows, 1):
        write_data_row(ws, i + 2, [row.get(h, "") for h in headers], i % 2 == 0)

    autofit(ws)
    return ws


# ── summary sheet ─────────────────────────────────────────────────────────────
def build_summary_sheet(wb, mock, extra):
    ws = wb.active
    ws.title = "المؤشر العام"
    set_sheet_style(ws)

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20

    title = "المؤشر العام - وزارة البلديات والإسكان | 2026"
    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value     = title
    c.font      = TITLE_FONT
    c.alignment = RTL_CENTER
    c.fill      = WHITE_FILL
    ws.row_dimensions[1].height = 36

    # section header helper
    def sec_header(row, label):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        c = ws.cell(row=row, column=1, value=label)
        c.font      = SUMM_HDR_FONT
        c.fill      = HDR_FILL
        c.alignment = RTL_CENTER
        ws.row_dimensions[row].height = 22

    def kv(row, label, value, alt=False):
        fill = ALT_FILL if alt else WHITE_FILL
        ws.row_dimensions[row].height = 18
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = Font(name="Arial", bold=True, size=10)
        lc.fill = fill; lc.alignment = RTL_RIGHT; lc.border = CELL_BORDER
        vc = ws.cell(row=row, column=2, value=value)
        vc.font = DATA_FONT; vc.fill = fill
        vc.alignment = RTL_CENTER; vc.border = CELL_BORDER
        # merge B+C for the value cell
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)

    r = 2

    # ── MOCK_DATA stats ──
    bi  = mock["bi"];  ops = mock["ops"]; strat = mock["strategy"]

    def count_status(lst, field, val):
        return sum(1 for x in lst if x.get(field,"") == val)

    def avg_completion(lst):
        vals = [x.get("نسبة الإنجاز الفعلي", 0) for x in lst if isinstance(x.get("نسبة الإنجاز الفعلي"), (int,float))]
        return round(sum(vals)/len(vals), 1) if vals else 0

    sec_header(r, "ملخص بيانات مركز ذكاء الأعمال")
    r += 1
    kv(r,   "إجمالي المشاريع",             len(bi),                        False); r += 1
    kv(r,   "مشاريع قيد التنفيذ",           count_status(bi,"تحديث الحالة التعاقدية","قيد التنفيذ"), True);  r += 1
    kv(r,   "مشاريع مكتملة",               count_status(bi,"تحديث الحالة التعاقدية","مكتمل"),       False); r += 1
    kv(r,   "مشاريع متأخرة",               count_status(bi,"تحديث الحالة التعاقدية","متأخر"),       True);  r += 1
    kv(r,   "متوسط نسبة الإنجاز (%)",       avg_completion(bi),             False); r += 1

    r += 1
    sec_header(r, "ملخص بيانات الأداء التشغيلي")
    r += 1
    kv(r,   "إجمالي البنود",               len(ops),                        False); r += 1
    kv(r,   "بنود مكتملة",                 count_status(ops,"الحالة","مكتمل"),     True);  r += 1
    kv(r,   "بنود على المسار",              count_status(ops,"الحالة","على المسار"),False); r += 1
    kv(r,   "بنود متأخرة",                 count_status(ops,"الحالة","متأخر"),     True);  r += 1
    kv(r,   "وثائق معتمدة",               count_status(ops,"حالة الوثائق","معتمدة"), False); r += 1

    r += 1
    sec_header(r, "ملخص بيانات التخطيط الاستراتيجي")
    r += 1
    kv(r,   "إجمالي المشاريع",             len(strat),                      False); r += 1
    kv(r,   "مشاريع قيد التنفيذ",           count_status(strat,"تحديث الحالة التعاقدية","قيد التنفيذ"), True);  r += 1
    kv(r,   "مشاريع مكتملة",               count_status(strat,"تحديث الحالة التعاقدية","مكتمل"),       False); r += 1
    kv(r,   "مشاريع متأخرة",               count_status(strat,"تحديث الحالة التعاقدية","متأخر"),       True);  r += 1
    kv(r,   "متوسط نسبة الإنجاز (%)",       avg_completion(strat),          False); r += 1

    r += 1
    sec_header(r, "ملخص البيانات الإضافية")
    r += 1
    kv(r, "عدد مؤشرات الأداء (بطاقة الأداء)",   len(extra["performance"]),   False); r += 1
    kv(r, "عدد مهام خطة العمل",                  len(extra["workplan"]),      True);  r += 1
    kv(r, "عدد التدخلات المسجلة",                 len(extra["interventions"]), False); r += 1
    kv(r, "عدد التقارير",                         len(extra["reports"]),       True);  r += 1

    # grand total
    r += 1
    total = len(bi) + len(ops) + len(strat)
    ws.row_dimensions[r].height = 22
    tc = ws.cell(row=r, column=1, value="إجمالي السجلات الكلي (MOCK_DATA)")
    tc.font = Font(name="Arial", bold=True, size=11, color=WHITE)
    tc.fill = HDR_FILL; tc.alignment = RTL_RIGHT; tc.border = CELL_BORDER
    vc = ws.cell(row=r, column=2, value=total)
    vc.font = Font(name="Arial", bold=True, size=11, color=WHITE)
    vc.fill = HDR_FILL; vc.alignment = RTL_CENTER; vc.border = CELL_BORDER
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)

    return ws


# ── main ──────────────────────────────────────────────────────────────────────
def main():
    js_path  = r"C:\Users\ZahraAAlhiyai\MOMAH-KPIs2026\data\mockData.js"
    out_dir  = r"C:\Users\ZahraAAlhiyai\MOMAH-KPIs2026\exports"
    out_path = os.path.join(out_dir, "MOMAH_KPIs_2026.xlsx")

    os.makedirs(out_dir, exist_ok=True)

    with open(js_path, encoding="utf-8") as f:
        js_text = f.read()

    mock, extra = extract_js_objects(js_text)

    wb = Workbook()

    build_summary_sheet(wb, mock, extra)

    add_data_sheet(wb, "ذكاء الأعمال",         mock["bi"])
    add_data_sheet(wb, "الأداء التشغيلي",       mock["ops"])
    add_data_sheet(wb, "التخطيط الاستراتيجي",   mock["strategy"])
    add_data_sheet(wb, "بطاقة الأداء",          extra["performance"])
    add_data_sheet(wb, "خطة العمل",             extra["workplan"])
    add_data_sheet(wb, "التدخلات",              extra["interventions"])
    add_data_sheet(wb, "التقارير",              extra["reports"])

    wb.save(out_path)
    print(f"Saved: {out_path}")

    # quick verification
    from openpyxl import load_workbook
    wb2 = load_workbook(out_path)
    for name in wb2.sheetnames:
        ws = wb2[name]
        print(f"  Sheet '{name}': {ws.max_row} rows x {ws.max_column} cols")


if __name__ == "__main__":
    main()
